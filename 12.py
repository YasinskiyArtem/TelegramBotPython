# ______________________________________________________________________________________________________________________
# |                                  CPO Cyber programming organization                                                |
# |____________________________________________________________________________________________________________________|
# |        Project Manager:             by https://vk.com/wolfnort                                                     |
# |____________________________________________________________________________________________________________________|
# |        core programmers:            by https://vk.com/ologn                                                        |
# |                                     by https://vk.com/y_a_n_1_x                                                    }
# |        documentation and testing:   by https://vk.com/id330572938                                                  }
# |                                     by https://vk.com/andpodryv                                                    |
# |                                     by https://vk.com/id114486343                                                  |
# |____________________________________________________________________________________________________________________|
# |        especially for: ROBBO Club.                                                                                 |
# |                        Ave Medikov, 5 building 7, St. Petersburg, 197022                                           |
# |____________________________________________________________________________________________________________________|

import openpyxl  # библиотека для работы с гугл таблицами
import telebot  # библиотека для работы с телеграм
from telebot import types  # из библиотеки telebot импортируем типы

from datetime import datetime  # библиотека для получения даты и времени
from keyboa import Keyboa  # библиотека для правильной работы клавиатуры
import threading  # библиотека для распараллеливания (один поток - один ученик)

# ____________________________________________________________________________________ #
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseDownload,MediaFileUpload
from googleapiclient.discovery import build
import pprint
from threading import Timer
from time import sleep
import time
# ВЫГРУЗКА В ДИСК: http://datalytics.ru/all/rabotaem-s-api-google-drive-s-pomoschyu-python/ #
# ____________________________________________________________________________________ #
pp = pprint.PrettyPrinter(indent=4)
SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'C:/Users/artem\PycharmProjects/untitled6/round-music-314016-2aa4ff2bd404.json'
credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)

current_datetime = datetime.now().time()
TIME_GOOGLE = str(current_datetime.hour) + ':' + str(current_datetime.minute)             # 12
print(TIME_GOOGLE)
print(current_datetime.minute - 8)
timing = TIME_GOOGLE




#while False:
 #   print('я тут1')
  # if (current_datetime.minute - 13 == 0):
   #     print('я тут1')
    #    service = build('drive', 'v3', credentials=credentials)
    #    folder_id = '1gpa9lQ8ji_Zr_vdWws_C1HgBwkmJBzsw'
     #   name = 'Результаты по городам.xlsx'
    #    name1 = 'Результаты по тестам.xlsx'
     #   file_path = 'C:/Users/artem/PycharmProjects/untitled6/Результаты по городам.xlsx'
     #   file_metadata = {
     #                   'name': name,
     #                   'parents': [folder_id]
    #                }
     #   media = MediaFileUpload(file_path, resumable=True)
     #   r = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
     #   pp.pprint(r)
     #   folder_id1 = '1gpa9lQ8ji_Zr_vdWws_C1HgBwkmJBzsw'
     #   name1 = 'Результаты по тестам.xlsx'
     #   file_path1 = 'C:/Users/artem/PycharmProjects/untitled6/Результаты по тестам.xlsx'
     #   file_metadata1 = {
     #                   'name': name1,
     #                   'parents': [folder_id1]
      #              }
     #   media1 = MediaFileUpload(file_path1, resumable=True)
     #   r = service.files().create(body=file_metadata1, media_body=media1, fields='id').execute()
     #   pp.pprint(r)


# ____________________________________________________________________________________________________________________ #
wbTest = openpyxl.reader.excel.load_workbook(filename="Тесты.xlsx")  # загрузка таблиц
wbClubResult = openpyxl.reader.excel.load_workbook(filename="Результаты по городам.xlsx")
wbTestResult = openpyxl.reader.excel.load_workbook(filename="Результаты по тестам.xlsx")
clubs = ["Анапа",
         "Астрахань",
         "Барнаул",
         "Бишкек",
         "Воронеж",
         "Екатеринбург",
         "Ижевск",
         "Калуга",
         "Керчь Врошилов",
         "Керчь Ленина",
         "Ковров",
         "Краснодар Монтажников",
         "Краснодар Московская",
         "Крымск",
         "ЛО Гатчина",
         "ЛО Никольское",
         "ЛО Новоселье",
         "ЛО Сосновый бор",
         "Магадан",
         "Москва Бибирево",
         "Москва ВДНХ",
         "Москва Зорге",
         "Москва Люберцы",
         "Москва Митино",
         "Москва Раменки",
         "МО Химки",
         "Мурманск Полярные Зори",
         "Мурманск Щербкова Нижний Новгород",
         "Оренбург",
         "Петропаловск-Камчатский",
         "Псков",
         "СПб Аристо",
         "СПб Беговая",
         "СПб Блюхера",
         "СПб Васильеостровский остров",
         "СПб Гете-Шуле",
         "СПб Гимназия 70",
         "СПб Ириновский",
         "СПб Комендантский",
         "СПб Косыгина",
         "СПб Лицей 82",
         "Санкт-Обводный",
         "СПб Петроградская",
         "СПб Политехническая",
         "СПб Шуваловский",
         "Северодвинск",
         "Серпухов",
         "Симферополь",
         "Славянск-на-Кубани",
         "Ставрополь",
         "Сургут",
         "Тамбов",
         "Темрюк"]  # все клубы которые существуют в РОББО

bot = telebot.TeleBot('1768613839:AAGjTnmo2SkX9R_s4Qccvu2pCFAzjd-iILQ')  # ключ бота для работы python- telegram

# Настройка нормального размера кнопок бота для каждой нашей клавиатуры
keyboardBack = telebot.types.ReplyKeyboardMarkup(True, True)
keyboardStart = telebot.types.ReplyKeyboardMarkup(True, True)
keyboardSubject = telebot.types.ReplyKeyboardMarkup(True, True)
keyboardYear = telebot.types.ReplyKeyboardMarkup(True, True)
keyboardRepeat = telebot.types.ReplyKeyboardMarkup(True, True)
hideMarkup = telebot.types.ReplyKeyboardRemove()
keyboardRepeat.row('ПОВТОРИТЬ ТЕСТ', 'ЗАВЕРШИТЬ')
keyboardBack.row('НАЗАД')
keyboardStart.row('НАЗАД', 'НАЧАТЬ')
keyboardYear.add('НАЗАД', '1 год (8-10)', '1 год (11-14)', '2 год (8-10)', '2 год (11-14)', '3 год (8-10)',
                 '3 год (11-14)')

# Cловарь позволяющий намм перепрыгивать в нужные ячейка, чтобы щаписать ФИ, количество баллов за предмет и послед сдачу
offset = {"1 год (8-10)": {'последняя сдача': 'B', 'фамилия': 'C', 'имя': 'D', 'Скретч': 'E', 'Лаборатория': 'F',
                           'Робот': 'G', 'Схемотехника 3D': 'H', '3D': 'I', 'Механика': 'J'},
          "1 год (11-14)": {'последняя сдача': 'M', 'фамилия': 'N', 'имя': 'O', 'Скретч': 'P', 'Лаборатория': 'Q',
                            'Робот': 'R', 'Схемотехника 3D': 'S', '3D': 'T', 'Механика': 'U'},
          "2 год (8-10)": {'последняя сдача': 'X', 'фамилия': 'Y', 'имя': 'Z', 'АппИнвентор': 'AA', 'Лаборатория': 'AB',
                           'Робот': 'AC', 'Схемотехника 3D': 'AD', '3D': 'AE'},
          "2 год (11-14)": {'последняя сдача': 'AH', 'фамилия': 'AI', 'имя': 'AJ', 'АппИнвентор': 'AK',
                            'Лаборатория': 'AL', 'Робот': 'AM', 'Схемотехника 3D': 'AN', '3D': 'AO'},
          "3 год (8-10)": {'последняя сдача': 'AR', 'фамилия': 'AS', 'имя': 'AT', 'Юнити': 'AU',
                           'Ардуино': 'AV', 'Интернет вещей': 'AW'},
          "3 год (11-14)": {'последняя сдача': 'AZ', 'фамилия': 'BA', 'имя': 'BB', 'Юнити': 'BC',
                            'Ардуино': 'BD', 'Интернет вещей': 'BE'}}
keyboardSubject.add('НАЗАД', 'Скретч', 'Лаборатория', 'Робот', 'Схемотехника 3D', '3D', 'Механика', 'АппИнвентор',
                    'Схемотехника 3D', 'Юнити', 'Ардуино', 'Интернет вещей')

# Подключаем наши кнопки клубов,так же прописываем функцию вывода в  2 строки

keyboardClubs = Keyboa(items=clubs, copy_text_to_callback=True, items_in_row=2).keyboard


@bot.message_handler(content_types=['text'])  # тип сообщений ,воспринимаемый ботом --- текст
def getTextMessages(message):
    # Сравнение с командой /test
    if message.text == '/test':
        x = threading.Thread(target=getStartClub, args=(message,))  # создать поток и запустить его
        x.start()
        y = threading.Thread(target=FILE_GOOGLE, args =(message,))
        y.start()

# функция выбора клуба. Бот передает сообщение с текстом + клавиатурой РОББОклубов
def getStartClub(message):
    bot.send_message(message.from_user.id, "Выберите свой РОББО Клуб", reply_markup=keyboardClubs)


@bot.callback_query_handler(func=lambda call: True)  # Словить нажатие клавиши и запомнить в club
def answerClub(query):
    bot.send_message(query.message.chat.id, query.data)
    bot.edit_message_reply_markup(query.message.chat.id, query.message.message_id)
    club = query.data
    getStartSubject(query.message, club)


def getStartSubject(message, club):  # сообщение о выборе предмета
    bot.send_message(message.chat.id, "Выберите предмет", reply_markup=keyboardSubject)
    bot.register_next_step_handler(message, getSubject, club)


def getSubject(message, club):
    subject = message.text  # запомнить предмет в subject
    if subject == 'НАЗАД':  # возврат при нажатии кнопки "НАЗАД"
        getStartClub(message)
    else:
        getStartSurname(message, club, subject)


def getStartSurname(message, club, subject):  # сообщение о выборе фамилии
    bot.send_message(message.from_user.id, "Введите свою фамилию", reply_markup=keyboardBack)
    bot.register_next_step_handler(message, getSurname, club, subject)


def getSurname(message, club, subject):
    if message.text == 'НАЗАД':
        getStartSubject(message, club)
    else:
        surname = message.text  # запомнить фамилию в surname
        getStartName(message, club, subject, surname)


def getStartName(message, club, subject, surname):  # сообщение о выборе имени
    bot.send_message(message.from_user.id, "Введите своё имя", reply_markup=keyboardBack)
    bot.register_next_step_handler(message, getName, club, subject, surname)


def getName(message, club, subject, surname):
    if message.text == 'НАЗАД':
        getStartSurname(message, club, subject)
    else:
        name = message.text  # запомнить имя в name
        getStartAge(message, club, subject, surname, name)


def getStartAge(message, club, subject, surname, name):  # сообщение о выборе возраста
    bot.send_message(message.from_user.id, "Введите свой возраст", reply_markup=keyboardBack)
    bot.register_next_step_handler(message, getAge, club, subject, surname, name)


def getAge(message, club, subject, surname, name):
    if message.text == 'НАЗАД':
        getStartName(message, club, subject, surname)
    else:
        age = message.text  # запоминание возраста в age
        if not age.isdigit():  # проверка на запись цифрами
            bot.send_message(message.from_user.id, 'Цифрами, пожалуйста')
            getStartAge(message, club, subject, surname, name)
        else:
            getStartYear(message, club, subject, surname, name, age)


def getStartYear(message, club, subject, surname, name, age):  # сообщение о выборе года обучения
    bot.send_message(message.from_user.id, "Выберите год обучения", reply_markup=keyboardYear)
    bot.register_next_step_handler(message, getYear, club, subject, surname, name, age)


def getYear(message, club, subject, surname, name, age):
    if message.text == 'НАЗАД':
        getStartAge(message, club, subject, surname, name)
    else:
        year = message.text  # запоминание года обучения в year
        getStartTeacher(message, club, subject, surname, name, age, year)


def getStartTeacher(message, club, subject, surname, name, age, year):  # сообщение о выборе педагога
    bot.send_message(message.from_user.id, "Введите фамилию и имя Вашего педагога", reply_markup=keyboardBack)
    bot.register_next_step_handler(message, getTeacher, club, subject, surname, name, age, year)


def getTeacher(message, club, subject, surname, name, age, year):
    if message.text == 'НАЗАД':
        getStartYear(message, club, subject, surname, name, age)
    else:
        teacher = message.text  # запоминание педагога в teacher
        bot.send_message(message.from_user.id, "Начать тест?", reply_markup=keyboardStart)  # сообщение о начале теста
        bot.register_next_step_handler(message, intermediaryFunc, club, subject, surname, name, age, year, teacher)


def intermediaryFunc(message, club, subject, surname, name, age, year, teacher):
    if message.text == 'НАЗАД':
        getStartTeacher(message, club, subject, surname, name, age, year)
    elif message.text == 'НАЧАТЬ':
        try:  # проверка на существование страницы с названием клуба в книге Результаты по городам
            wbClubResult[club]
        except KeyError:
            wbClubResult.create_sheet(club)

        # Приведение формата имени, фамилии, и имени педагога к одному формату (начало каждого слова с большой буквы,
        # вместо ё - е)
        if str(name).endswith('\n') or str(name).endswith(' '):
            name = name[:-1]
        name.replace('ё', 'е')
        name.replace('Ё', 'Е')
        name = str(name).title()
        if str(surname).endswith('\n') or str(surname).endswith(' '):
            surname = surname[:-1]
        surname.replace('ё', 'е')
        surname.replace('Ё', 'Е')
        surname = str(surname).title()
        if str(teacher).endswith('\n') or str(teacher).endswith(' '):
            teacher = teacher[:-1]
        teacher.replace('ё', 'е')
        teacher.replace('Ё', 'Е')
        teacher = str(teacher).title()

        sheetClubResult = wbClubResult[str(club)]  # получение страницы в Результаты по городам с названием
        # выбранного клуба
        j = 3
        # поиск свободного места и одновремеенно проверка на наличие данных фамилии и имени в книге
        # Результаты по городам на странице с названием выбранного клуба
        while sheetClubResult[offset[year]['фамилия'] + str(j)].value is not None:
            if sheetClubResult[offset[year]['фамилия'] + str(j)].value == surname and sheetClubResult[
                offset[year]['имя'] + str(j)].value == name:
                break
            else:
                j += 1
        freeSpace = j
        # запись в найденную ячейку имени и фамилии
        sheetClubResult[offset[year]['фамилия'] + str(freeSpace)].value = surname
        sheetClubResult[offset[year]['имя'] + str(freeSpace)].value = name
        intermediaryFunc2(message, club, subject, surname, name, age, year, teacher, freeSpace)


def intermediaryFunc2(message, club, subject, surname, name, age, year, teacher, freeSpace):
    # открытие в книге Результаты по тестам страницы с названием предмета и годом обучения
    sheetTestResult = wbTestResult[subject + ' ' + year]
    # заполнение начальных клеток
    sheetTestResult['A' + str(1)] = "Отметка времени"
    sheetTestResult['B' + str(1)] = "Баллы"
    sheetTestResult['C' + str(1)] = "Фамилия учащегося"
    sheetTestResult['D' + str(1)] = "Имя учащегося"
    sheetTestResult['E' + str(1)] = "Возраст"
    sheetTestResult['F' + str(1)] = "РОББО Клуб"
    sheetTestResult['G' + str(1)] = "Фамилия и имя педагога"
    j = 1
    # поиск свободной ячейки в книге Результаты по тестам страницы с названием предмета и годом обучения
    while sheetTestResult['C' + str(j)].value is not None:
        j += 1
    freeSpaceTest = j
    # запись данных об учение
    sheetTestResult['C' + str(freeSpaceTest)].value = surname
    sheetTestResult['D' + str(freeSpaceTest)].value = name
    sheetTestResult['E' + str(freeSpaceTest)].value = age
    sheetTestResult['F' + str(freeSpaceTest)].value = club
    sheetTestResult['G' + str(freeSpaceTest)].value = teacher
    i = 2
    score = 0
    maxScore = 0
    exitTest = 0
    # название листа с тестом, состоящее из предмета и годом обучения
    testSheet = str(subject) + ' ' + str(year)
    startTest(message, i, score, maxScore, exitTest, club, subject, surname, name, age, year, teacher, freeSpace,
              freeSpaceTest,
              testSheet)


def startTest(message, i, score, maxScore, exitTest, club, subject, surname, name, age, year, teacher, freeSpace,
              freeSpaceTest,
              testSheet):
    try:  # Попытка открыть лист с тестом
        wbTest[testSheet]
    except KeyError:
        # если его нет, возврат на пункт выбора клуба (самое начало)
        bot.send_message(message.from_user.id, 'Тест пуст')
        getStartClub(message)
    else:
        sheetTest = wbTest[testSheet]
        if sheetTest['A' + str(1)].value is None:
            # если лист пуст, возврат на пункт выбора клуба (самое начало)
            bot.send_message(message.from_user.id, 'Тест пуст')
            getStartClub(message)
        else:  # начало теста
            if message.text != '/test' and message.text != 'НАЧАТЬ':
                if str(sheetTest['K' + str(i - 1)].value).isdigit():  # подсчёт максимального числа баллов из столбца K
                    maxScore += int(sheetTest['K' + str(i - 1)].value)
                else:
                    # если ячейка пуста - вес вопроса 1 балл
                    maxScore += 1
                if message.text == sheetTest['J' + str(i - 1)].value:
                    # если ответ совпал с правильным из столбца J - добавляется баллы равные весу вопроса
                    score += sheetTest['K' + str(i - 1)].value
                # открытие в книге Результаты по тестам листа с названием состоящим из
                # названия предмета и годом обучения
                sheetTestResult = wbTestResult[subject + ' ' + year]
                k = chr(ord('E') + i)
                # запись ответа ученика
                sheetTestResult[k + str(freeSpaceTest)] = message.text
            sheetTest = wbTest[testSheet]
            if exitTest == 1:  # метка выхода из теста. Ставится на последнем вопросе
                finishTest(message, i, score, maxScore, exitTest, club, subject, surname, name, age, year, teacher,
                           freeSpace,
                           freeSpaceTest, testSheet)
            else:
                # создание клавиатуры с вариантами ответа, если такие имеются
                keyboardAnswer = types.ReplyKeyboardMarkup(True, one_time_keyboard=True)
                for k in ('B', 'C', 'D', 'E', 'F', 'G', 'I'):
                    if sheetTest[k + str(i)].value is None:
                        break
                    else:
                        answer = str(sheetTest[k + str(i)].value)
                        keyboardAnswer.add(answer)
                # получение вопроса со страницы с тестами
                question = str(sheetTest['A' + str(i)].value)
                # получение всех встроенных в вопрос картинок согласно формату
                ind = question.find("\\картинка")
                if ind != -1:
                    paths = question[ind + 10:]
                    if paths.endswith('\n'):
                        paths = paths[:-1]
                    lst = paths.split(' ')
                    question = question[:ind]
                    # отправка вопроса
                    bot.send_message(message.from_user.id, question)
                    for path in lst:
                        # отправка картинок
                        bot.send_photo(message.from_user.id, open(str(path), 'rb'), reply_markup=keyboardAnswer)
                else:  # отправка вопроса
                    bot.send_message(message.from_user.id, question, reply_markup=keyboardAnswer)
                # если вопрос последний, ставится метка выхода из теста
                if sheetTest['A' + str(i + 1)].value is None:
                    exitTest = 1
                i += 1
                bot.register_next_step_handler(message, startTest, i, score, maxScore, exitTest, club, subject, surname,
                                               name,
                                               age, year, teacher, freeSpace, freeSpaceTest, testSheet)


def finishTest(message, i, score, maxScore, exitTest, club, subject, surname, name, age, year, teacher, freeSpace,
               freeSpaceTest, testSheet):
    # открытие нужных листов из книги с результатами по городам
    sheetClubResult = wbClubResult[club]
    # и книги с результатами по тестам
    sheetTestResult = wbTestResult[subject + ' ' + year]
    # Запись результата в книгу Результаты по тестам
    sheetTestResult['B' + str(freeSpaceTest)] = str(score) + "\\" + str(maxScore)
    # Если в результатах по городам уже есть данный ученик и результат по тесту, новый записывается рядом
    if sheetClubResult[offset[year][subject] + str(freeSpace)].value is not None:
        sheetClubResult[offset[year][subject] + str(freeSpace)] = sheetClubResult[offset[year][subject] + str(
            freeSpace)].value + ' ' + str(score) + "\\" + str(maxScore)
    # иначе заполняет клетку
    else:
        sheetClubResult[offset[year][subject] + str(freeSpace)] = str(score) + "\\" + str(maxScore)
    # получение метки времени
    sheetTestResult['A' + str(freeSpaceTest)] = sheetClubResult[offset[year]['последняя сдача'] + str(freeSpace)] = \
        str(datetime.now().date()) + ' ' + str(datetime.now().hour) + ':' + str(datetime.now().minute) + ':' + str(
            datetime.now().second)
    # сохранение полученных записей в книги
    wbClubResult.save('Результаты по городам.xlsx')
    wbTestResult.save('Результаты по тестам.xlsx')
    # сообщение об окончании теста и предложение пройти ещё раз
    bot.send_message(message.from_user.id, "Поздравляем! Вы успешно прошли тест", reply_markup=keyboardRepeat)
    bot.register_next_step_handler(message, repeatFunc, i, score, maxScore, exitTest, club, subject, surname, name, age,
                                   year, teacher, freeSpace,
                                   freeSpaceTest, testSheet)



def repeatFunc(message, i, score, maxScore, exitTest, club, subject, surname, name, age, year, teacher, freeSpace,
               freeSpaceTest, testSheet):
    # возможность повторить тест
    if message.text == 'ПОВТОРИТЬ ТЕСТ':
        i = 2
        score = 0
        maxScore = 0
        exitTest = 0
        startTest(message, i, score, maxScore, exitTest, club, subject, surname, name, age, year, teacher, freeSpace,
                  freeSpaceTest, testSheet)


# ____________________________________________________________________________________________________________________ #
# ЗАПИСЬ В ГУГЛ: http://datalytics.ru/all/rabotaem-s-api-google-drive-s-pomoschyu-python/                              #
# ____________________________________________________________________________________________________________________ #
def first_file(): # запись первого файла
        sleep(3600) # ожидание 1 час, можно поменять на 12 часов
        service = build('drive', 'v3', credentials=credentials)
        folder_id = '1gpa9lQ8ji_Zr_vdWws_C1HgBwkmJBzsw'
        name = 'Результаты по городам.xlsx'
        name1 = 'Результаты по тестам.xlsx'
        file_path = 'C:/Users/artem/PycharmProjects/untitled6/Результаты по городам.xlsx' # ВАШ ПУТЬ
        file_metadata = {
                        'name': name,
                        'parents': [folder_id]
                    }
        media = MediaFileUpload(file_path, resumable=True)
        r = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        pp.pprint(r)

def next_file(): # fаналогично с выше
        sleep(3600)
        service = build('drive', 'v3', credentials=credentials)
        folder_id1 = '1gpa9lQ8ji_Zr_vdWws_C1HgBwkmJBzsw'
        name1 = 'Результаты по тестам.xlsx'
        file_path1 = 'C:/Users/artem/PycharmProjects/untitled6/Результаты по тестам.xlsx' # ВАШ ПУТЬ
        file_metadata1 = {
                        'name': name1,
                        'parents': [folder_id1]
                    }
        media1 = MediaFileUpload(file_path1, resumable=True)
        r = service.files().create(body=file_metadata1, media_body=media1, fields='id').execute()
        pp.pprint(r)

def FILE_GOOGLE(message): # функция- таймер
    while 1:
        timer = Timer(5, first_file)
        timer.start()
        next_file()

bot.polling(none_stop=True, interval=0)